"""Tests for cleanings calendar helpers in cleanings/main.py."""

from __future__ import annotations

import json
import shutil
from datetime import date
from pathlib import Path

import openpyxl
import pytest

from cleanings.main import (
    booking_listing_label,
    build_cleaning_sheet,
    contiguous_column_runs,
    day_rows_by_month,
    outgoing_same_day_checkout_codes,
    start_column_for_month,
    stay_days_by_month,
    stay_guest_label,
    year_from_bookings,
    FIRST_DAY_COL,
    TEMPLATE_WEEKDAY_PHASE,
)
from shared.listing_labels import LISTING_LABELS

T2 = "Totalmente Renovado, metro à porta"


def _booking(
    listing_name: str,
    start_date: str,
    end_date: str,
    *,
    confirmation_code: str = "HMTEST001",
    adults: int = 2,
) -> dict:
    return {
        "confirmationCode": confirmation_code,
        "listingName": listing_name,
        "startDate": start_date,
        "endDate": end_date,
        "numberOfAdults": adults,
        "numberOfChildren": 0,
        "numberOfInfants": 0,
    }


class TestYearFromBookings:
    def test_reads_year_from_date_range(self) -> None:
        data = {"dateRange": {"startDate": "2026-01-01", "endDate": "2026-12-31"}}
        assert year_from_bookings(data) == 2026


class TestBookingListingLabel:
    def test_maps_known_listing_name(self) -> None:
        booking = _booking(T2, "2026-06-01", "2026-06-05")
        assert booking_listing_label(booking) == LISTING_LABELS[T2]

    def test_returns_none_for_unknown_listing(self) -> None:
        booking = _booking("Unknown listing", "2026-06-01", "2026-06-05")
        assert booking_listing_label(booking) is None


class TestStayGuestLabel:
    def test_formats_adults_children_infants(self) -> None:
        booking = {
            "numberOfAdults": 2,
            "numberOfChildren": 1,
            "numberOfInfants": 1,
        }
        assert stay_guest_label(booking) == "2a 1c 1b"

    def test_omits_zero_counts(self) -> None:
        assert stay_guest_label({"numberOfAdults": 1}) == "1a"


class TestStartColumnForMonth:
    def test_january_2026_starts_on_thursday_column(self) -> None:
        # 2026-01-01 is Thursday; template phase 0 => column D is Monday
        assert start_column_for_month(2026, 1, TEMPLATE_WEEKDAY_PHASE) == FIRST_DAY_COL + 3


class TestContiguousColumnRuns:
    def test_groups_consecutive_columns(self) -> None:
        assert contiguous_column_runs([4, 5, 6, 9, 10]) == [(4, 6), (9, 10)]

    def test_empty_list(self) -> None:
        assert contiguous_column_runs([]) == []


class TestStayDaysByMonth:
    def test_splits_stay_across_month_boundary(self) -> None:
        start = date(2026, 1, 30)
        end = date(2026, 2, 2)
        by_month = stay_days_by_month(start, end, 2026)
        assert by_month[1] == [date(2026, 1, 30), date(2026, 1, 31)]
        assert by_month[2] == [date(2026, 2, 1), date(2026, 2, 2)]


class TestOutgoingSameDayCheckoutCodes:
    def test_skips_outgoing_when_checkout_equals_next_checkin(self) -> None:
        bookings = [
            _booking(T2, "2026-06-01", "2026-06-05", confirmation_code="OUT"),
            _booking(T2, "2026-06-05", "2026-06-10", confirmation_code="IN"),
        ]
        assert outgoing_same_day_checkout_codes(bookings) == {"OUT"}


class TestBuildCleaningSheet:
    def test_writes_year_sheet_from_template(self, tmp_path: Path) -> None:
        repo_root = Path(__file__).resolve().parents[2]
        template = repo_root / "cleanings" / "templates" / "cleanings-map.xlsx"
        workbook = tmp_path / "cleanings-map.xlsx"
        shutil.copy(template, workbook)

        bookings_path = tmp_path / "bookings.json"
        bookings_path.write_text(
            json.dumps(
                {
                    "dateRange": {
                        "startDate": "2026-01-01",
                        "endDate": "2026-12-31",
                    },
                    "bookings": [
                        _booking(T2, "2026-03-10", "2026-03-14"),
                    ],
                }
            ),
            encoding="utf-8",
        )

        year, saved_path, html_path = build_cleaning_sheet(bookings_path, workbook)
        assert year == 2026
        assert saved_path == workbook
        assert html_path == tmp_path / "cleanings-2026.html"
        assert html_path.is_file()
        assert "2026" in html_path.read_text(encoding="utf-8")

        wb = openpyxl.load_workbook(workbook)
        assert "2026" in wb.sheetnames
        ws = wb["2026"]
        jan_day_row = day_rows_by_month(ws)["Janeiro"]
        jan_start_col = start_column_for_month(2026, 1, TEMPLATE_WEEKDAY_PHASE)
        assert ws.cell(jan_day_row, jan_start_col).value == 1

    def test_skips_html_when_disabled(self, tmp_path: Path) -> None:
        repo_root = Path(__file__).resolve().parents[2]
        template = repo_root / "cleanings" / "templates" / "cleanings-map.xlsx"
        workbook = tmp_path / "cleanings-map.xlsx"
        shutil.copy(template, workbook)

        bookings_path = tmp_path / "bookings.json"
        bookings_path.write_text(
            json.dumps(
                {
                    "dateRange": {
                        "startDate": "2026-01-01",
                        "endDate": "2026-12-31",
                    },
                    "bookings": [],
                }
            ),
            encoding="utf-8",
        )

        _, _, html_path = build_cleaning_sheet(
            bookings_path, workbook, write_html=False
        )
        assert html_path is None
        assert not (tmp_path / "cleanings-2026.html").exists()

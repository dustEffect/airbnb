#!/usr/bin/env python3
"""Build the annual cleaning calendar sheet from bookings.json."""

from __future__ import annotations

import argparse
import calendar
import json
from copy import copy
from datetime import date, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from fetch.run_fetch import add_fetch_arguments, maybe_run_fetch
from shared.listing_labels import LISTING_LABELS
from shared.paths import bookings_json_path, project_root

CLEANINGS_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = project_root()
DEFAULT_BOOKINGS = bookings_json_path()
DEFAULT_WORKBOOK = CLEANINGS_DIR / "templates" / "cleanings-map.xlsx"
TEMPLATE_SHEET = "template"
FIRST_DAY_COL = 4  # D
LAST_DAY_COL = 40  # AN — one past AM so long months align with the template weekday row
TEMPLATE_WEEKDAY_PHASE = 0  # column D = Monday, matching the template row
WEEKEND_FILL = PatternFill(patternType="solid", fgColor="FFA5A5A5")
NO_FILL = PatternFill()
THIN_SIDE = Side(style="thin")
ROWS_PER_MONTH_BLOCK = 6

MONTHS_PT = (
    "Janeiro",
    "Fevereiro",
    "Março",
    "Abril",
    "Maio",
    "Junho",
    "Julho",
    "Agosto",
    "Setembro",
    "Outubro",
    "Novembro",
    "Dezembro",
)


def year_from_bookings(data: dict) -> int:
    return int(data["dateRange"]["startDate"][:4])


def day_rows_by_month(ws: Worksheet) -> dict[str, int]:
    rows: dict[str, int] = {}
    for row in range(1, ws.max_row + 1):
        month = ws.cell(row, 2).value
        label = ws.cell(row, 3).value
        if month in MONTHS_PT and label == "day":
            rows[month] = row
    missing = [month for month in MONTHS_PT if month not in rows]
    if missing:
        raise ValueError(f"Missing day rows for months: {', '.join(missing)}")
    return rows


def start_column_for_month(year: int, month: int, phase: int) -> int:
    first_weekday = date(year, month, 1).weekday()
    for col in range(FIRST_DAY_COL, LAST_DAY_COL + 1):
        if (col - FIRST_DAY_COL + phase) % 7 == first_weekday:
            return col
    raise ValueError(f"No weekday column found for {year}-{month:02d}")


def is_weekend_column(col: int, phase: int) -> bool:
    return (col - FIRST_DAY_COL + phase) % 7 in (5, 6)


def apply_weekend_fills(ws: Worksheet, phase: int) -> None:
    day_rows = day_rows_by_month(ws)
    for month_name in MONTHS_PT:
        day_row = day_rows[month_name]
        for row in range(day_row, day_row + ROWS_PER_MONTH_BLOCK):
            for col in range(FIRST_DAY_COL, LAST_DAY_COL + 1):
                cell = ws.cell(row, col)
                has_day = ws.cell(day_row, col).value is not None
                if has_day and is_weekend_column(col, phase):
                    cell.fill = WEEKEND_FILL
                else:
                    cell.fill = NO_FILL


def fill_day_row(ws: Worksheet, day_row: int, year: int, month: int, phase: int) -> None:
    num_days = calendar.monthrange(year, month)[1]
    start_col = start_column_for_month(year, month, phase)
    end_col = start_col + num_days - 1
    if end_col > LAST_DAY_COL:
        raise ValueError(
            f"{MONTHS_PT[month - 1]} {year} needs column "
            f"{openpyxl.utils.get_column_letter(end_col)} but grid ends at AM"
        )

    for col in range(FIRST_DAY_COL, LAST_DAY_COL + 1):
        ws.cell(day_row, col).value = None

    for day in range(1, num_days + 1):
        ws.cell(day_row, start_col + day - 1, day)


def listing_rows_in_month(ws: Worksheet, day_row: int) -> dict[str, int]:
    rows: dict[str, int] = {}
    for offset in range(1, ROWS_PER_MONTH_BLOCK):
        row = day_row + offset
        label = ws.cell(row, 3).value
        if label:
            rows[str(label)] = row
    return rows


def column_for_day(ws: Worksheet, day_row: int, day: int) -> int | None:
    for col in range(FIRST_DAY_COL, LAST_DAY_COL + 1):
        if ws.cell(day_row, col).value == day:
            return col
    return None


def booking_listing_label(booking: dict) -> str | None:
    listing_name = booking.get("listingName")
    if not listing_name:
        return None
    return LISTING_LABELS.get(listing_name.strip())


def stay_guest_label(booking: dict) -> str:
    parts: list[str] = []
    adults = booking.get("numberOfAdults", 0)
    children = booking.get("numberOfChildren", 0)
    infants = booking.get("numberOfInfants", 0)
    if adults:
        parts.append(f"{adults}a")
    if children:
        parts.append(f"{children}c")
    if infants:
        parts.append(f"{infants}b")
    return " ".join(parts)


def apply_stay_cell_alignment(cell) -> None:
    alignment = copy(cell.alignment)
    alignment.horizontal = "left"
    cell.alignment = alignment


def apply_stay_cell_font(cell) -> None:
    font = copy(cell.font)
    font.color = "FF000000"
    font.bold = True
    cell.font = font


def contiguous_column_runs(columns: list[int]) -> list[tuple[int, int]]:
    if not columns:
        return []

    runs: list[tuple[int, int]] = []
    start = previous = columns[0]
    for column in columns[1:]:
        if column == previous + 1:
            previous = column
            continue
        runs.append((start, previous))
        start = previous = column
    runs.append((start, previous))
    return runs


def remove_internal_vertical_borders(ws: Worksheet, row: int, columns: list[int]) -> None:
    for start_col, end_col in contiguous_column_runs(columns):
        if start_col == end_col:
            continue
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row, col)
            border = copy(cell.border)
            if col > start_col:
                border.left = Side(style=None)
            if col < end_col:
                border.right = Side(style=None)
            cell.border = border


def smooth_stay_borders(ws: Worksheet) -> None:
    day_rows = day_rows_by_month(ws)
    for month_name in MONTHS_PT:
        day_row = day_rows[month_name]
        for listing_row in listing_rows_in_month(ws, day_row).values():
            listing_color = ws.cell(listing_row, 3).fill.fgColor.rgb
            painted_cols = [
                col
                for col in range(FIRST_DAY_COL, LAST_DAY_COL + 1)
                if ws.cell(listing_row, col).fill.fgColor.rgb == listing_color
            ]
            remove_internal_vertical_borders(ws, listing_row, painted_cols)


def stay_days_by_month(start: date, end: date, year: int) -> dict[int, list[date]]:
    days_by_month: dict[int, list[date]] = {}
    current = start
    while current <= end:
        if current.year == year:
            days_by_month.setdefault(current.month, []).append(current)
        current += timedelta(days=1)
    return days_by_month


def smooth_cross_month_stay_borders(
    ws: Worksheet, year: int, bookings: list[dict]
) -> None:
    day_rows = day_rows_by_month(ws)
    for booking in bookings:
        label = booking_listing_label(booking)
        if not label:
            continue

        start = date.fromisoformat(booking["startDate"])
        end = date.fromisoformat(booking["endDate"])
        days_by_month = stay_days_by_month(start, end, year)
        months = sorted(days_by_month)
        if len(months) < 2:
            continue

        for month_index in range(len(months) - 1):
            month_a = months[month_index]
            month_b = months[month_index + 1]
            if month_b != month_a + 1:
                continue

            last_day_a = max(days_by_month[month_a])
            first_day_b = min(days_by_month[month_b])
            if first_day_b != last_day_a + timedelta(days=1):
                continue

            day_row_a = day_rows[MONTHS_PT[month_a - 1]]
            day_row_b = day_rows[MONTHS_PT[month_b - 1]]
            listing_row_a = listing_rows_in_month(ws, day_row_a).get(label)
            listing_row_b = listing_rows_in_month(ws, day_row_b).get(label)
            if listing_row_a is None or listing_row_b is None:
                continue

            col_a = column_for_day(ws, day_row_a, last_day_a.day)
            col_b = column_for_day(ws, day_row_b, first_day_b.day)
            if col_a is None or col_b is None:
                continue

            border_a = copy(ws.cell(listing_row_a, col_a).border)
            border_a.right = Side(style=None)
            ws.cell(listing_row_a, col_a).border = border_a

            border_b = copy(ws.cell(listing_row_b, col_b).border)
            border_b.left = Side(style=None)
            ws.cell(listing_row_b, col_b).border = border_b


def outgoing_same_day_checkout_codes(bookings: list[dict]) -> set[str]:
    """Confirmation codes for outgoing stays that share checkout/check-in day."""
    skipped_codes: set[str] = set()
    by_label: dict[str, list[dict]] = {}

    for booking in bookings:
        label = booking_listing_label(booking)
        if not label:
            continue
        by_label.setdefault(label, []).append(booking)

    for listing_bookings in by_label.values():
        listing_bookings.sort(key=lambda booking: booking["startDate"])
        for index in range(len(listing_bookings) - 1):
            checkout = date.fromisoformat(listing_bookings[index]["endDate"])
            checkin = date.fromisoformat(listing_bookings[index + 1]["startDate"])
            if checkout == checkin:
                code = listing_bookings[index].get("confirmationCode")
                if code:
                    skipped_codes.add(code)

    return skipped_codes


def bookings_by_label(bookings: list[dict]) -> dict[str, list[dict]]:
    by_label: dict[str, list[dict]] = {}
    for booking in bookings:
        label = booking_listing_label(booking)
        if not label:
            continue
        by_label.setdefault(label, []).append(booking)
    for listing_bookings in by_label.values():
        listing_bookings.sort(key=lambda booking: booking["startDate"])
    return by_label


def set_cell_border_side(
    ws: Worksheet,
    listing_row: int,
    col: int,
    *,
    left: bool = False,
    right: bool = False,
) -> None:
    border = copy(ws.cell(listing_row, col).border)
    if left:
        border.left = THIN_SIDE
    if right:
        border.right = THIN_SIDE
    ws.cell(listing_row, col).border = border


def apply_stay_boundary_borders(
    ws: Worksheet, year: int, bookings: list[dict]
) -> None:
    day_rows = day_rows_by_month(ws)

    for booking in bookings:
        label = booking_listing_label(booking)
        if not label:
            continue

        start = date.fromisoformat(booking["startDate"])
        if start.year != year:
            continue

        month_name = MONTHS_PT[start.month - 1]
        day_row = day_rows[month_name]
        listing_row = listing_rows_in_month(ws, day_row).get(label)
        checkin_col = column_for_day(ws, day_row, start.day)
        if listing_row is None or checkin_col is None:
            continue

        set_cell_border_side(ws, listing_row, checkin_col, left=True)

    for listing_bookings in bookings_by_label(bookings).values():
        for index in range(len(listing_bookings) - 1):
            outgoing = listing_bookings[index]
            incoming = listing_bookings[index + 1]
            label = booking_listing_label(outgoing)
            if not label:
                continue

            checkout = date.fromisoformat(outgoing["endDate"])
            checkin = date.fromisoformat(incoming["startDate"])

            if checkout == checkin:
                checkout_day = checkout - timedelta(days=1)
                checkin_day = checkin
            elif checkin == checkout + timedelta(days=1):
                checkout_day = checkout
                checkin_day = checkin
            else:
                continue

            for boundary_day, side in (
                (checkout_day, "right"),
                (checkin_day, "left"),
            ):
                if boundary_day.year != year:
                    continue
                month_name = MONTHS_PT[boundary_day.month - 1]
                day_row = day_rows[month_name]
                listing_row = listing_rows_in_month(ws, day_row).get(label)
                col = column_for_day(ws, day_row, boundary_day.day)
                if listing_row is None or col is None:
                    continue
                set_cell_border_side(
                    ws,
                    listing_row,
                    col,
                    left=side == "left",
                    right=side == "right",
                )


def apply_booking_stays(ws: Worksheet, year: int, bookings: list[dict]) -> None:
    day_rows = day_rows_by_month(ws)
    skipped_outgoing_codes = outgoing_same_day_checkout_codes(bookings)
    for booking in bookings:
        label = booking_listing_label(booking)
        if not label:
            continue

        start = date.fromisoformat(booking["startDate"])
        end = date.fromisoformat(booking["endDate"])
        guest_label = stay_guest_label(booking)
        current = start
        while current <= end:
            if (
                booking.get("confirmationCode") in skipped_outgoing_codes
                and current == end
            ):
                current += timedelta(days=1)
                continue
            if current.year == year:
                month_name = MONTHS_PT[current.month - 1]
                day_row = day_rows[month_name]
                listing_row = listing_rows_in_month(ws, day_row).get(label)
                if listing_row is None:
                    raise ValueError(
                        f"Listing row '{label}' not found for {month_name}"
                    )

                col = column_for_day(ws, day_row, current.day)
                if col is not None:
                    cell = ws.cell(listing_row, col)
                    cell.fill = copy(ws.cell(listing_row, 3).fill)
                    apply_stay_cell_alignment(cell)
                    apply_stay_cell_font(cell)
                    if current == start and guest_label:
                        cell.value = guest_label

            current += timedelta(days=1)


def fill_year_days(ws: Worksheet, year: int) -> None:
    day_rows = day_rows_by_month(ws)
    for month, month_name in enumerate(MONTHS_PT, start=1):
        fill_day_row(
            ws,
            day_rows[month_name],
            year,
            month,
            TEMPLATE_WEEKDAY_PHASE,
        )
    apply_weekend_fills(ws, TEMPLATE_WEEKDAY_PHASE)


def clone_template_sheet(wb: openpyxl.Workbook, year: int) -> Worksheet:
    sheet_name = str(year)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    if TEMPLATE_SHEET not in wb.sheetnames:
        raise ValueError(f"Workbook is missing '{TEMPLATE_SHEET}' sheet")

    template = wb[TEMPLATE_SHEET]
    ws = wb.copy_worksheet(template)
    ws.title = sheet_name
    return ws


def build_cleaning_sheet(
    bookings_path: Path = DEFAULT_BOOKINGS,
    workbook_path: Path = DEFAULT_WORKBOOK,
) -> tuple[int, Path]:
    data = json.loads(bookings_path.read_text(encoding="utf-8"))
    year = year_from_bookings(data)

    wb = openpyxl.load_workbook(workbook_path)
    ws = clone_template_sheet(wb, year)
    fill_year_days(ws, year)
    bookings = data.get("bookings", [])
    apply_booking_stays(ws, year, bookings)
    smooth_stay_borders(ws)
    smooth_cross_month_stay_borders(ws, year, bookings)
    apply_stay_boundary_borders(ws, year, bookings)
    wb.save(workbook_path)
    return year, workbook_path


def main() -> None:
    parser = argparse.ArgumentParser(
        prog="python cleanings/main.py",
        description=(
            "Clone the cleaning calendar template for the booking year, fill day "
            "rows, highlight weekends, and paint stay windows from shared/bookings.json.\n\n"
            "Fetches shared/bookings.json from Airbnb by default for the full "
            "calendar year. Pass --no-fetch to reuse an existing file."
        ),
    )
    add_fetch_arguments(parser, offer_no_fetch=True, include_range_args=False)
    parser.add_argument(
        "--year",
        type=int,
        default=date.today().year,
        metavar="YYYY",
        help="Calendar year to fetch from Airbnb (default: current year).",
    )
    parser.add_argument(
        "--bookings",
        type=Path,
        default=DEFAULT_BOOKINGS,
        help=f"Path to bookings.json (default: {DEFAULT_BOOKINGS.relative_to(PROJECT_ROOT)})",
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        default=DEFAULT_WORKBOOK,
        help=f"Path to cleanings-map.xlsx (default: {DEFAULT_WORKBOOK.relative_to(PROJECT_ROOT)})",
    )
    args = parser.parse_args()

    maybe_run_fetch(args, calendar_year=None if args.no_fetch else args.year)

    if not args.bookings.is_file():
        raise SystemExit(
            f"Missing {args.bookings}. Run without --no-fetch or run ./fetch.sh."
        )

    year, workbook_path = build_cleaning_sheet(args.bookings, args.workbook)
    print(f"Updated {workbook_path} with sheet '{year}'")


if __name__ == "__main__":
    main()
